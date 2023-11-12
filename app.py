from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory
from flask_wtf import FlaskForm
from wtforms import StringField
from wtforms.validators import DataRequired
from werkzeug.utils import secure_filename
import os
import openpyxl
from flask import jsonify

app = Flask(__name__)
app.config['SECRET_KEY'] = 'amaljyothi884'
app.config['UPLOAD_FOLDER'] = 'uploads'

class UploadForm(FlaskForm):
    file = StringField('Excel File', validators=[DataRequired()])

@app.route('/')
def index():
  
    return render_template('home.html')

@app.route('/display', methods=['GET', 'POST'])
def upload_file():
    form = UploadForm()

    if request.method == 'POST' and form.validate_on_submit():
        file = request.files['file']
        if file:
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            flash('File uploaded successfully!', 'success')
            return redirect(url_for('display_data', filename=filename))

    return render_template('index.html', form=form)

@app.route('/display/<filename>', methods=['GET', 'POST'])
def display_data(filename):
    form = UploadForm()
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    if not os.path.exists(file_path):
        flash('File not found!', 'danger')
        return redirect(url_for('upload_file'))

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]

    if request.method == 'POST':
        selected_row = int(request.form.get('selected_row'))
        data = [cell.value for cell in sheet[selected_row]]
        return render_template('index.html', headers=headers, data=data, selected_row=selected_row, filename=filename)

    return render_template('index.html', headers=headers, sheet=sheet, filename=filename,form=form)

@app.route('/update/<filename>/<int:selected_row>', methods=['POST'])
def update_data(filename, selected_row):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    if not os.path.exists(file_path):
        flash('File not found!', 'danger')
        return redirect(url_for('upload_file'))

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Update the data in the workbook based on the AJAX request
    for i, field in enumerate(request.form.values()):
        sheet.cell(row=selected_row, column=i + 1, value=field)

    workbook.save(file_path)

    # Return a JSON response to indicate success
    return jsonify({'status': 'success'})

@app.route('/download/<filename>')
def download_data(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

if __name__ == '__main__':
    if not os.path.exists(app.config['UPLOAD_FOLDER']):
        os.makedirs(app.config['UPLOAD_FOLDER'])
    app.run(debug=True)
